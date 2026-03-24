"""
data_server.py  —  Run this on the RECEIVING laptop alongside mqtt_logger.py.

Reads bms_data.xlsx and exposes it as a JSON API so the HTML dashboard
can poll it live.

Install once:
    pip install flask openpyxl flask-cors

Run:
    python data_server.py

Then open index.html in a browser on the same machine.
"""

from flask import Flask, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook
import os

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")

EXCEL_FILE  = "bms_data.xlsx"
HISTORY_MAX = 20   # rows shown in the history table

# Column order matches mqtt_logger.py headers:
# Timestamp, Topic, Room, CO2, Temperature, Humidity, Raw Payload
COL = {
    "timestamp":   0,
    "topic":       1,
    "room":        2,
    "co2":         3,
    "temperature": 4,
    "humidity":    5,
    "raw":         6,
}


def read_rows():
    if not os.path.exists(EXCEL_FILE):
        return []
    try:
        wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:      # skip header
                continue
            if any(v is not None for v in row):
                rows.append(list(row))
        wb.close()
        return rows
    except Exception as exc:
        print(f"[server] Excel read error: {exc}")
        return []


def row_to_dict(row):
    def val(key):
        v = row[COL[key]] if COL[key] < len(row) else None
        if v is None:
            return ""
        return str(v) if not isinstance(v, (int, float)) else v

    return {
        "timestamp":   val("timestamp"),
        "topic":       val("topic"),
        "room":        val("room"),
        "co2":         val("co2"),
        "temperature": val("temperature"),
        "humidity":    val("humidity"),
    }


@app.route("/data")
def data():
    rows = read_rows()
    if not rows:
        return jsonify({"latest": None, "history": [], "total": 0})

    latest_row  = rows[-1]
    recent_rows = list(reversed(rows[-HISTORY_MAX:]))

    return jsonify({
        "latest":  row_to_dict(latest_row),
        "history": [row_to_dict(r) for r in recent_rows],
        "total":   len(rows),
    })


@app.route("/health")
def health():
    exists = os.path.exists(EXCEL_FILE)
    rows   = read_rows() if exists else []
    return jsonify({
        "status":       "ok",
        "excel_found":  exists,
        "record_count": len(rows),
    })


if __name__ == "__main__":
    print("=" * 50)
    print("  BMS Data Server — localhost:5000")
    print(f"  Excel file : {EXCEL_FILE}")
    print("  Endpoints  : /data   /health")
    print("=" * 50)
    app.run(host="0.0.0.0", port=5000, debug=False)
