"""
dashboard.py  —  Run this on the RECEIVING laptop (or any laptop that can
                 see  bms_data.xlsx).

Reads the Excel file every few seconds and prints a live-updating terminal
dashboard showing the latest readings and a recent history table.

Install deps once:
    pip install openpyxl
"""

import os
import time
from datetime import datetime
from openpyxl import load_workbook

# ── Configuration ─────────────────────────────────────────────────────────────
EXCEL_FILE      = "bms_data.xlsx"   # must be in the same folder
SHEET_NAME      = "Sensor Log"
REFRESH_SECONDS = 3                 # how often to re-read the file
HISTORY_ROWS    = 10                # rows shown in the "recent history" table

# Column positions in the sheet (1-based, matching mqtt_logger.py headers)
COL_TIMESTAMP   = 1
COL_TOPIC       = 2
COL_ROOM        = 3
COL_CO2         = 4
COL_TEMP        = 5
COL_HUMIDITY    = 6
COL_RAW         = 7

# ── Terminal colours (ANSI) ───────────────────────────────────────────────────
RESET  = "\033[0m"
BOLD   = "\033[1m"
CYAN   = "\033[96m"
GREEN  = "\033[92m"
YELLOW = "\033[93m"
RED    = "\033[91m"
BLUE   = "\033[94m"
DIM    = "\033[2m"


def clear():
    os.system("cls" if os.name == "nt" else "clear")


def read_excel_rows():
    """Return all data rows (skip header) as list-of-lists, newest first."""
    if not os.path.exists(EXCEL_FILE):
        return []
    try:
        wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:          # skip header
                continue
            if any(cell is not None for cell in row):
                rows.append(list(row))
        wb.close()
        return list(reversed(rows))   # newest first
    except Exception as exc:
        return []


def co2_bar(value):
    """Return a coloured bar representing CO2 level."""
    try:
        v = int(value)
    except (TypeError, ValueError):
        return DIM + "N/A" + RESET

    bar_len = min(int(v / 100), 20)
    bar     = "█" * bar_len + "░" * (20 - bar_len)

    if v < 600:
        colour = GREEN
        label  = "Good"
    elif v < 1000:
        colour = YELLOW
        label  = "Moderate"
    else:
        colour = RED
        label  = "High"

    return f"{colour}{bar}{RESET}  {colour}{BOLD}{v} ppm  [{label}]{RESET}"


def fmt(value, unit="", fallback="—"):
    if value is None or value == "":
        return DIM + fallback + RESET
    return f"{BOLD}{value}{RESET}{unit}"


def draw_dashboard(rows):
    latest = rows[0] if rows else None
    total  = len(rows)
    now    = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")

    clear()
    width = 70
    print(CYAN + "═" * width + RESET)
    print(CYAN + BOLD +
          "  🏢  BMS REAL-TIME SENSOR DASHBOARD".center(width) + RESET)
    print(CYAN + f"  Last updated: {now}".ljust(width) + RESET)
    print(CYAN + "═" * width + RESET)

    if latest is None:
        print(f"\n  {YELLOW}⏳  Waiting for data in {EXCEL_FILE} …{RESET}\n")
        return

    print(f"\n  {BOLD}{'LATEST READING':}{RESET}")
    print(f"  {'─'*60}")
    print(f"  {BLUE}Timestamp :{RESET}  {fmt(latest[COL_TIMESTAMP - 1])}")
    print(f"  {BLUE}Topic     :{RESET}  {fmt(latest[COL_TOPIC - 1])}")
    print(f"  {BLUE}Room      :{RESET}  {fmt(latest[COL_ROOM - 1])}")
    print(f"  {BLUE}CO2       :{RESET}  {co2_bar(latest[COL_CO2 - 1])}")
    print(f"  {BLUE}Temp      :{RESET}  {fmt(latest[COL_TEMP - 1], ' °C')}")
    print(f"  {BLUE}Humidity  :{RESET}  {fmt(latest[COL_HUMIDITY - 1], ' %')}")
    print(f"  {'─'*60}")
    print(f"  {DIM}Total records stored: {total}{RESET}\n")

    # ── Recent history table ──────────────────────────────────────────────────
    recent = rows[:HISTORY_ROWS]
    print(f"  {BOLD}RECENT HISTORY  (last {len(recent)} readings){RESET}")
    print(f"  {'─'*60}")
    hdr = (f"  {'Timestamp':<22} {'Room':<14} {'CO2':>8}  "
           f"{'Temp':>6}  {'Hum':>5}")
    print(BLUE + BOLD + hdr + RESET)
    print(f"  {'─'*60}")

    for r in recent:
        ts   = str(r[COL_TIMESTAMP - 1] or "")[-19:]
        room = str(r[COL_ROOM - 1] or "")[:13]
        co2  = str(r[COL_CO2 - 1] or "—")
        temp = str(r[COL_TEMP - 1] or "—")
        hum  = str(r[COL_HUMIDITY - 1] or "—")

        # Colour CO2 value
        try:
            co2_int = int(co2)
            co2_col = GREEN if co2_int < 600 else (YELLOW if co2_int < 1000 else RED)
        except ValueError:
            co2_col = RESET

        line = (f"  {ts:<22} {room:<14} "
                f"{co2_col}{co2:>8}{RESET}  "
                f"{temp:>6}  {hum:>5}")
        print(line)

    print(f"\n  {DIM}Refreshing every {REFRESH_SECONDS}s  —  Ctrl-C to quit{RESET}")
    print(CYAN + "═" * width + RESET)


# ── Main loop ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"Starting dashboard — watching {EXCEL_FILE}  …")
    time.sleep(1)

    try:
        while True:
            rows = read_excel_rows()
            draw_dashboard(rows)
            time.sleep(REFRESH_SECONDS)
    except KeyboardInterrupt:
        clear()
        print(f"\n{YELLOW}Dashboard stopped.{RESET}\n")
