"""
mqtt_logger.py  —  Run this on the RECEIVING laptop (where Mosquitto is running).

It subscribes to all bms/# topics and appends every incoming message
to  bms_data.xlsx  in the same folder.

Install deps once:
    pip install paho-mqtt openpyxl
"""

import json
import os
import time
from datetime import datetime

import paho.mqtt.client as mqtt
from paho.mqtt.enums import CallbackAPIVersion
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configuration ────────────────────────────────────────────────────────────
BROKER_HOST   = "localhost"   # Mosquitto is running on THIS machine
BROKER_PORT   = 1883
SUBSCRIBE_TOPIC = "bms/#"    # Wildcard — catches every bms/... topic
EXCEL_FILE    = "bms_data.xlsx"
SHEET_NAME    = "Sensor Log"
JS_FILE       = "sensor_data.js"
HISTORY_MAX   = 20

HEADERS = ["Timestamp", "Topic", "Room", "CO2 (ppm)", "Temperature (°C)",
           "Humidity (%)", "Raw Payload"]

# ── Excel helpers ─────────────────────────────────────────────────────────────

def _style_header_row(ws):
    header_fill  = PatternFill("solid", start_color="1F4E79")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center       = Alignment(horizontal="center", vertical="center")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center

    col_widths = [22, 28, 16, 14, 18, 16, 50]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def get_or_create_workbook():
    """Load existing workbook or create a fresh one with a styled header."""
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        _style_header_row(ws)
        wb.save(EXCEL_FILE)
        print(f"[logger] Created {EXCEL_FILE}")
    return wb, ws


def append_row(payload_dict: dict, topic: str, raw: str):
    """Append one row to the Excel sheet and save."""
    wb, ws = get_or_create_workbook()

    readings = payload_dict.get("readings", {})
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        topic,
        payload_dict.get("room", ""),
        readings.get("co2", ""),
        readings.get("temperature", ""),
        readings.get("humidity", ""),
        raw,
    ]

    # Alternate row shading for readability
    next_row = ws.max_row + 1
    fill_color = "D6E4F0" if next_row % 2 == 0 else "FFFFFF"
    row_fill = PatternFill("solid", start_color=fill_color)
    row_font = Font(name="Arial", size=10)

    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.fill = row_fill
        cell.font = row_font
        cell.alignment = Alignment(horizontal="center")

    wb.save(EXCEL_FILE)



# ── JS data file writer ───────────────────────────────────────────────────────

def write_js_file(payload_dict: dict, topic: str, timestamp: str):
    """Read last HISTORY_MAX rows from Excel and write sensor_data.js."""
    try:
        wb, ws = get_or_create_workbook()
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if any(v is not None for v in row):
                rows.append(row)
        wb.close()
    except Exception:
        rows = []

    history = []
    for row in reversed(rows[-HISTORY_MAX:]):
        history.append({
            "timestamp":   str(row[0]) if row[0] else "",
            "topic":       str(row[1]) if row[1] else "",
            "room":        str(row[2]) if row[2] else "",
            "co2":         row[3] if row[3] != "" and row[3] is not None else "",
            "temperature": row[4] if row[4] != "" and row[4] is not None else "",
            "humidity":    row[5] if row[5] != "" and row[5] is not None else "",
        })

    readings = payload_dict.get("readings", {})
    latest = {
        "timestamp":   timestamp,
        "topic":       topic,
        "room":        payload_dict.get("room", ""),
        "co2":         readings.get("co2", ""),
        "temperature": readings.get("temperature", ""),
        "humidity":    readings.get("humidity", ""),
    }

    data = {"latest": latest, "history": history, "total": len(rows)}
    js_content = f"window.SENSOR_DATA = {json.dumps(data, indent=2)};"

    with open(JS_FILE, "w") as f:
        f.write(js_content)


# ── MQTT callbacks ────────────────────────────────────────────────────────────

def on_connect(client, userdata, flags, reason_code, properties):
    if reason_code == 0:
        print(f"[logger] ✅ Connected to Mosquitto on {BROKER_HOST}:{BROKER_PORT}")
        client.subscribe(SUBSCRIBE_TOPIC)
        print(f"[logger] 📡 Subscribed to  '{SUBSCRIBE_TOPIC}'")
    else:
        print(f"[logger] ❌ Connection refused — reason code {reason_code}")


def on_message(client, userdata, msg):
    raw = msg.payload.decode("utf-8", errors="replace")
    ts  = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] ← {msg.topic}  |  {raw}")

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        data = {"room": "unknown", "readings": {}}

    try:
        append_row(data, msg.topic, raw)
        write_js_file(data, msg.topic, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    except Exception as exc:
        print(f"[logger] ⚠️  Excel write error: {exc}")


def on_disconnect(client, userdata, flags, reason_code, properties):
    print(f"[logger] 🔌 Disconnected (code {reason_code}). Reconnecting in 5 s…")
    time.sleep(5)


# ── Main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 55)
    print("  BMS MQTT → Excel Logger")
    print(f"  Broker  : {BROKER_HOST}:{BROKER_PORT}")
    print(f"  Topic   : {SUBSCRIBE_TOPIC}")
    print(f"  Output  : {EXCEL_FILE}")
    print("=" * 55)

    # Ensure file exists before first message arrives
    get_or_create_workbook()

    mqttc = mqtt.Client(CallbackAPIVersion.VERSION2, "BMS_Logger")
    mqttc.on_connect    = on_connect
    mqttc.on_message    = on_message
    mqttc.on_disconnect = on_disconnect

    mqttc.connect(BROKER_HOST, BROKER_PORT, keepalive=60)
    mqttc.loop_forever()   # blocks; Ctrl-C to stop
